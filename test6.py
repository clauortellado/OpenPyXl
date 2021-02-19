# Add a Sheet to the Existing XLSX

from openpyxl import load_workbook

filepath="C:/Users/Klau/Documents/Python/XLS/"
file1 = "demo3.xlsx"
file2 = "demo4.xlsx"
wb = load_workbook(filepath+file1)
sheet = wb.active

#create a new sheet
wb.create_sheet('Sheet2')
wb.save(filepath+file2)
