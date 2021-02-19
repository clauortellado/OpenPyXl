# Reading from Excel Sheet

from openpyxl import load_workbook

filepath="C:/Users/Klau/Documents/Python/XLS/"
file1 = "demo3.xlsx"
wb = load_workbook(filepath+file1)
sheet = wb.active

b1 = sheet['A1']
b2 = sheet['B2']
b3 = sheet.cell(row=3, column=3)

print(b1.value)
print(b2.value)
print(b3.value)
