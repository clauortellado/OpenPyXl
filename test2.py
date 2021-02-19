# https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
# Working with Excel sheets in Python using openpyxl

# Writing to a cell____________________________________
from openpyxl import load_workbook

filepath = "C:/Users/Klau/Documents/Python/XLS"
file1 = "demo1.xlsx"
file2 = "demo2.xlsx"

wb= load_workbook(filepath+"/"+file1)
sheet=wb.active

sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2

wb.save(filepath+"/"+file2)
