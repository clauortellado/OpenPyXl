# Iterating by Rows

from openpyxl import load_workbook

filepath="C:/Users/Klau/Documents/Python/XLS/"
file1 = "demo3.xlsx"
wb = load_workbook(filepath+file1)
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

#Iterating over all cell - all rows
for i in range(1,max_row+1):
    #Iterating over all columns
    for j in range(1,max_col+1):
        #get particular cell value
        cell_obj=sheet.cell(row=i,column=j)
        #print cell value
        print(cell_obj.value,end=' | ')

    #print new line
    print('\n')

